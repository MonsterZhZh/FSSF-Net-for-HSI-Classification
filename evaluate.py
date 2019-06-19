#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Thu Dec  7 15:52:15 2017

@author: zpl
"""
import numpy as np
from sklearn.metrics import cohen_kappa_score
from sklearn.metrics import accuracy_score
from openpyxl import Workbook
import copy

def res2excel(config_name):
    # load result from npz file
    res = np.load('../result/res_{0:s}.npz'.format(config_name))
    res_predict_pretrain = res['RES_PREDICT_PRETRAIN']
    res_predict_finetune = res['RES_PREDICT_FINETUNE']
    res_gt = res['RES_GT']
    N_REPEAT = res_predict_finetune.shape[0]
    N_CLASS = res_predict_finetune.shape[2]
    
    # add oa,aa,kappa
    ct_pretrain = correct_table(res_predict_pretrain, res_gt)
    ct_finetune = correct_table(res_predict_finetune, res_gt)
    # write excel file
    wb = Workbook()
    ws = wb.active
    
    # write pretrain
    for i in range(N_CLASS+3):
        for j in range(N_REPEAT + 1):
            ws.cell(row=i+1, column=j+1, value=ct_pretrain[i][j])
            
    # write fintune
    for i in range(N_CLASS+3):
        for j in range(N_REPEAT + 1):
            ws.cell(row=i+1, column=j+1+N_REPEAT + 2, value=ct_finetune[i][j])
    # Save the file
    wb.save("../result/res_{0:s}.xlsx".format(config_name))
    return

def correct_table(res_predict, res_gt):
    res_max = res_predict.argmax(axis=2)
    res_correct = res_max == res_gt
    
    N_REPEAT = res_predict.shape[0]
    N_SAMPLE = res_predict.shape[1]
    N_CLASS = res_predict.shape[2]
    
    ct = np.zeros((N_CLASS, N_REPEAT + 1))
    for i in range(N_SAMPLE):
        ct[res_gt[0,i], 0] += 1
        for j in range(N_REPEAT):
            if res_correct[j,i]:
                ct[res_gt[j,i], j+1] += 1
    ct = ct.tolist()

    oa = overall_accuracy(res_predict, res_gt)
    aa = average_accuracy(res_predict, res_gt)
    kappa = kappa_accuracy(res_predict, res_gt)
    
    ct.append(oa)
    ct.append(aa)
    ct.append(kappa)
    return ct

def overall_accuracy(res_predict, res_gt):
    res_max = res_predict.argmax(axis=2)
    res_correct = res_max == res_gt
    sum_correct = res_correct.sum(axis = 1)
    oa = sum_correct / res_gt.shape[1]
    oa = oa.tolist()
    oa.insert(0,'OA')
    return oa

def average_accuracy(res_predict, res_gt):
    res_max = res_predict.argmax(axis=2)
    res_correct = res_max == res_gt
    
    N_REPEAT = res_predict.shape[0]
    N_SAMPLE = res_predict.shape[1]
    N_CLASS = res_predict.shape[2]
    aa = np.zeros((N_CLASS, N_REPEAT + 1))
    for i in range(N_SAMPLE):
        aa[res_gt[0,i], 0] += 1
        for j in range(N_REPEAT):
            if res_correct[j,i]:
                aa[res_gt[j,i], j+1] += 1
    aa = aa.transpose()
    aa = aa / aa[0]
    aa = aa[1:]
    aa = aa.mean(axis = 1)
    aa = aa.tolist()
    aa.insert(0,'AA')
    return aa

def kappa_accuracy(res_predict, res_gt):
    pred_label = res_predict.argmax(-1) # last axis 
    kappa = []
    for i in range(res_gt.shape[0]):
        kappa.append(cohen_kappa_score(pred_label[i,:],res_gt[i,:]))
    kappa.insert(0,'Kappa')
    return kappa
